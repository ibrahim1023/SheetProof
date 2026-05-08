import json
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Tax Rate"
    ws["B1"] = 0.2
    ws["C1"] = "=B1*100"
    wb.save(path)


def _run_audit(workdir: Path, workbook: Path) -> None:
    env = dict(os.environ)
    env["PYTHONPATH"] = str((Path(__file__).resolve().parents[2] / "src"))
    cmd = [sys.executable, "-m", "sheetproof.cli", "audit", str(workbook), "--deterministic"]
    subprocess.run(cmd, cwd=workdir, env=env, check=True, capture_output=True, text=True)


def test_reproducibility_across_fresh_processes(tmp_path: Path) -> None:
    workbook_path = tmp_path / "model.xlsx"
    _create_workbook(workbook_path)

    _run_audit(tmp_path, workbook_path)
    first_manifest = json.loads(
        (tmp_path / ".sheetproof" / "reproducibility-manifest.json").read_text(encoding="utf-8")
    )

    _run_audit(tmp_path, workbook_path)
    second_manifest = json.loads(
        (tmp_path / ".sheetproof" / "reproducibility-manifest.json").read_text(encoding="utf-8")
    )

    first_map = {x["file"]: x["canonical_sha256"] for x in first_manifest["artifacts"]}
    second_map = {x["file"]: x["canonical_sha256"] for x in second_manifest["artifacts"]}
    assert first_map == second_map
