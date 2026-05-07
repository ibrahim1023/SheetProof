from pathlib import Path

from openpyxl import Workbook

from sheetproof.assumptions.detector import detect_assumptions
from sheetproof.formulas.extractor import extract_formula_inventory
from sheetproof.graph.builder import build_dependency_graph
from sheetproof.graph.impact import compute_downstream_impact
from sheetproof.risk.rules import (
    detect_broken_reference_findings,
    detect_formula_inconsistency_findings,
    detect_hardcoded_override_findings,
    detect_hidden_external_dependency_findings,
    detect_volatile_formula_findings,
)
from sheetproof.risk.scorer import score_findings
from sheetproof.workbook.parser import parse_workbook


def test_detectors_and_scoring(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.12
    ws["B2"] = "=NOW()"
    ws["C2"] = "=B2+1"
    ws["D2"] = "=B2+2"
    ws["E2"] = 10
    ws["F2"] = "=#REF!+1"

    hidden = wb.create_sheet("HiddenSheet")
    hidden.sheet_state = "hidden"
    ws["A3"] = "=HiddenSheet!A1"

    p = tmp_path / "risk.xlsx"
    wb.save(p)

    index = parse_workbook(p)
    inventory = extract_formula_inventory(index)
    graph = build_dependency_graph(inventory)
    impact = compute_downstream_impact(graph)

    findings = []
    findings.extend(detect_formula_inconsistency_findings(inventory))
    findings.extend(detect_hardcoded_override_findings(index))
    findings.extend(detect_hidden_external_dependency_findings(index, inventory))
    findings.extend(detect_volatile_formula_findings(inventory))
    findings.extend(detect_broken_reference_findings(inventory))

    assert any(f.type == "hidden_sheet_dependency" for f in findings)
    assert any(f.type == "volatile_formula" for f in findings)
    assert any(f.type == "broken_reference" for f in findings)

    scored = score_findings(findings, impact)
    assert all(f.risk_score >= 0 for f in scored)


def test_assumption_detection(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Discount Rate"
    ws["B1"] = 0.08
    ws["C1"] = "=B1*100"

    p = tmp_path / "assumption.xlsx"
    wb.save(p)

    index = parse_workbook(p)
    inventory = extract_formula_inventory(index)
    graph = build_dependency_graph(inventory)
    impact = compute_downstream_impact(graph)

    assumptions = detect_assumptions(index, impact)
    assert any(a.label == "Discount Rate" for a in assumptions)
