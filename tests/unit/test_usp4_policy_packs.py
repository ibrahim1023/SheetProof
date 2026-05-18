import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app
from sheetproof.config.loader import load_config


runner = CliRunner()


def _workbook_with_volatile(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=NOW()"
    wb.save(path)


def test_policy_pack_changes_effective_policy_and_findings(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "p.xlsx"
    _workbook_with_volatile(workbook_path)

    monkeypatch.chdir(tmp_path)
    result_warn = runner.invoke(app, ["audit", str(workbook_path), "--policy-pack", "finance"])
    assert result_warn.exit_code == 0
    report_warn = json.loads(Path(".sheetproof/sheetproof-report.json").read_text(encoding="utf-8"))
    volatile_warn = [f for f in report_warn["findings"] if f["type"] == "volatile_formula"]
    assert volatile_warn

    result_allow = runner.invoke(app, ["audit", str(workbook_path), "--policy-pack", "operations"])
    assert result_allow.exit_code == 0
    report_allow = json.loads(Path(".sheetproof/sheetproof-report.json").read_text(encoding="utf-8"))
    assert "effective_policy" in report_allow


def test_invalid_schema_fails_closed(tmp_path: Path) -> None:
    bad = tmp_path / "bad.yml"
    bad.write_text("schema_version: 999\n", encoding="utf-8")

    try:
        load_config(config_path=bad)
        assert False, "Expected ValueError for invalid schema"
    except ValueError as exc:
        assert "schema_version" in str(exc)


def test_unknown_policy_pack_fails_closed(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "q.xlsx"
    _workbook_with_volatile(workbook_path)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(app, ["audit", str(workbook_path), "--policy-pack", "unknown-pack"])
    assert result.exit_code != 0


def test_policy_pack_metadata_present_and_non_empty() -> None:
    cfg = load_config()
    packs = cfg["policy_packs"]
    for pack_name in ("finance", "compliance", "operations"):
        metadata = packs[pack_name]["metadata"]
        assert metadata["version"]
        assert metadata["owner"]
        assert metadata["rationale"]
        assert metadata["updated_at"]
