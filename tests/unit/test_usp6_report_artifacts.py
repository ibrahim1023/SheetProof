import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Tax Rate"
    ws["B1"] = 0.2
    ws["C1"] = "=B1*100"
    wb.save(path)


def test_markdown_has_required_sections(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "r.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(app, ["audit", str(p), "--deterministic"])
    assert result.exit_code == 0

    md = Path(".sheetproof/sheetproof-report.md").read_text(encoding="utf-8")
    assert "## Human Review Checklist" in md
    assert "## Evidence Appendix" in md


def test_json_and_csv_artifacts_are_parseable_and_consistent(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "r2.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(app, ["audit", str(p), "--deterministic"])
    assert result.exit_code == 0

    report = json.loads(Path(".sheetproof/sheetproof-report.json").read_text(encoding="utf-8"))
    assert "workbook" in report
    assert "summary" in report
    assert "findings" in report
    assert "assumptions" in report

    risk_csv = Path(".sheetproof/risk-cells.csv").read_text(encoding="utf-8").splitlines()
    assumptions_csv = Path(".sheetproof/assumption-register.csv").read_text(encoding="utf-8").splitlines()
    assert len(risk_csv) >= 1
    assert len(assumptions_csv) >= 1
    assert risk_csv[0].startswith("Type,Severity,Sheet,Cell")
    coverage = json.loads(Path(".sheetproof/coverage-matrix.json").read_text(encoding="utf-8"))
    assert coverage["version"] == 1
    assert "warning_taxonomy" in coverage
