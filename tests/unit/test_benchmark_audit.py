import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Revenue"
    ws["B1"] = 100
    ws["A2"] = "Cost"
    ws["B2"] = 40
    ws["A3"] = "Margin"
    ws["B3"] = "=B1-B2"
    wb.save(path)


def test_benchmark_command_writes_output(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    wb = tmp_path / "bench.xlsx"
    _workbook(wb)
    out = tmp_path / "bench.json"
    res = runner.invoke(
        app,
        ["benchmark-audit", "--workbook", str(wb), "--runs", "2", "--output", str(out)],
    )
    assert res.exit_code == 0
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert payload["runs"] == 2
    assert payload["deterministic"] is True
    assert payload["runtime_ms"]["p95_ms"] >= 0


def test_benchmark_command_fails_regression_gate(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    wb = tmp_path / "bench.xlsx"
    _workbook(wb)
    baseline = tmp_path / "baseline.json"
    baseline.write_text(
        json.dumps(
            {
                "workbook": "bench.xlsx",
                "runs": 3,
                "deterministic": True,
                "runtime_ms": {"p95_ms": 0.0001},
                "workbook_stats": {},
            }
        ),
        encoding="utf-8",
    )
    res = runner.invoke(
        app,
        [
            "benchmark-audit",
            "--workbook",
            str(wb),
            "--runs",
            "2",
            "--baseline",
            str(baseline),
            "--max-regression-pct",
            "0.1",
        ],
    )
    assert res.exit_code == 22
