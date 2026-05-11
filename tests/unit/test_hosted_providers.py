import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    wb.save(path)


def _structured_json() -> str:
    return json.dumps(
        {
            "summary": "Cell C1 doubles growth input.",
            "risks": ["Input volatility"],
            "reviewer_actions": ["Verify Summary!B1"],
            "citations": [{"cell": "Summary!B1", "reason": "Direct dependency"}],
        }
    )


def test_openai_provider_path(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "o.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {"enabled": True, "provider": "openai", "model": "gpt-5.5"}
        },
    )
    monkeypatch.setattr("sheetproof.llm.local_explainer.call_openai", lambda **kwargs: _structured_json())

    result = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert result.exit_code == 0


def test_anthropic_provider_path(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "a.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {"enabled": True, "provider": "anthropic", "model": "claude-opus-4-1"}
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.call_anthropic", lambda **kwargs: _structured_json()
    )

    result = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert result.exit_code == 0


def test_gemini_provider_path(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "g.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {"llm": {"enabled": True, "provider": "gemini", "model": "gemini-2.5-pro"}},
    )
    monkeypatch.setattr("sheetproof.llm.local_explainer.call_gemini", lambda **kwargs: _structured_json())

    result = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert result.exit_code == 0
