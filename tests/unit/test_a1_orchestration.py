import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    wb.save(path)


def test_orchestration_traces_emitted(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "n.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {"enabled": True, "provider": "local", "model": "qwen", "base_url": "http://localhost:11434"}
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.explain_with_ollama",
        lambda prompt, model, base_url="http://localhost:11434": json.dumps(
            {
                "summary": "Valid summary text",
                "risks": [],
                "reviewer_actions": [],
                "citations": [],
            }
        ),
    )

    res = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert res.exit_code == 0
    traces = Path(".sheetproof/traces.jsonl").read_text(encoding="utf-8").splitlines()
    assert any('"event": "explain_start"' in line for line in traces)
    assert any('"event": "provider_call"' in line for line in traces)
    assert any('"event": "explain_success"' in line for line in traces)


def test_orchestration_accepts_fenced_json_output(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "n.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {"enabled": True, "provider": "local", "model": "qwen", "base_url": "http://localhost:11434"}
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.explain_with_ollama",
        lambda prompt, model, base_url="http://localhost:11434": """```json
{
  "summary": "Fenced JSON should be accepted by explain flow.",
  "risks": [],
  "reviewer_actions": [],
  "citations": []
}
```""",
    )

    res = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert res.exit_code == 0
    assert "Summary:" in res.stdout


def test_orchestration_coerces_variant_schema_output(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "n.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {"enabled": True, "provider": "local", "model": "qwen", "base_url": "http://localhost:11434"}
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.explain_with_ollama",
        lambda prompt, model, base_url="http://localhost:11434": """{
  "summary": {"cell": "Summary!C1", "text": "Cell C1 doubles B1."},
  "risks": [{"risk": "Input drift"}],
  "reviewer_actions": [{"action": "Verify Summary!B1 input owner."}],
  "citations": [{"cell": "Summary!B1", "reason": "Direct reference"}]
}""",
    )

    res = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert res.exit_code == 0
    assert "Summary: Cell C1 doubles B1." in res.stdout
