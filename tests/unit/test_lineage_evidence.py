import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    ws["D1"] = "=C1+1"
    ws["E1"] = "=D1+1"
    ws["F1"] = "=E1+1"
    ws["G1"] = "=[ext.xlsx]Sheet1!A1"

    hidden = wb.create_sheet("HiddenSheet")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 7
    ws["H1"] = "=HiddenSheet!A1"

    wb.save(path)


def test_findings_have_lineage_and_graph_consistency(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "lineage.xlsx"
    _build_workbook(workbook_path)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(app, ["audit", str(workbook_path), "--deterministic"])
    assert result.exit_code == 0

    report = json.loads(Path(".sheetproof/sheetproof-report.json").read_text(encoding="utf-8"))
    graph = json.loads(Path(".sheetproof/dependency-graph.json").read_text(encoding="utf-8"))
    graph_nodes = set(graph["nodes"])

    findings = report["findings"]
    assert findings, "Expected at least one finding for lineage validation"

    for finding in findings:
        assert finding["dependency_path"], f"Missing dependency_path for {finding['id']}"
        assert finding["source_cells"] is not None
        assert finding["path_depth"] >= 0
        for node in finding["dependency_path"]:
            assert node in graph_nodes


def test_risk_csv_contains_lineage_columns(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "lineage2.xlsx"
    _build_workbook(workbook_path)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(app, ["audit", str(workbook_path), "--deterministic"])
    assert result.exit_code == 0

    header = Path(".sheetproof/risk-cells.csv").read_text(encoding="utf-8").splitlines()[0]
    assert "DependencyPath" in header
    assert "SourceCells" in header
    assert "ImpactedOutputs" in header
    assert "PathDepth" in header
