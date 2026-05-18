from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app
from sheetproof.diff.workbook_diff import compute_workbook_diff


runner = CliRunner()


def _write_pair(tmp_path: Path) -> tuple[Path, Path]:
    old_wb = Workbook()
    old_ws = old_wb.active
    old_ws.title = "Inputs"
    old_ws["A1"] = "Growth Rate"
    old_ws["B1"] = 0.1
    old_ws["C1"] = "=B1*100"
    old_ws["D1"] = "=C1+5"

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Inputs"
    new_ws["A1"] = "Growth Rate"
    new_ws["B1"] = 0.12
    new_ws["C1"] = "=B1*120"
    new_ws["D1"] = "=C1+5"
    new_ws["E1"] = "=[ext.xlsx]Sheet1!A1"

    hidden = new_wb.create_sheet("HiddenCalc")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 1

    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    old_wb.save(old_path)
    new_wb.save(new_path)
    return old_path, new_path


def test_compute_workbook_diff(tmp_path: Path) -> None:
    old_path, new_path = _write_pair(tmp_path)
    result = compute_workbook_diff(old_path, new_path)

    assert result.changed_formulas >= 1
    assert result.changed_values >= 1
    assert result.changed_assumptions >= 1
    assert result.newly_hidden_sheets == ["HiddenCalc"]
    assert result.high_risk_changed_cells >= 1
    assert any(c.change_type == "formula_changed" for c in result.changes)
    assert isinstance(result.reviewer_summary, list)
    if result.reviewer_summary:
        assert "evidence_pointer" in result.reviewer_summary[0]


def test_cli_diff_writes_json(tmp_path: Path) -> None:
    old_path, new_path = _write_pair(tmp_path)
    result = runner.invoke(app, ["diff", str(old_path), str(new_path)])

    assert result.exit_code == 0
    assert "Changed formulas" in result.stdout
    assert "Workbook diff written" in result.stdout
    assert Path(".sheetproof/workbook-diff.json").exists()
