import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _create(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Revenue Growth"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    wb.save(path)


def test_findings_have_deterministic_provenance(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "w.xlsx"
    _create(p)
    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["audit", str(p), "--deterministic"])
    assert result.exit_code == 0

    report = json.loads(Path(".sheetproof/sheetproof-report.json").read_text(encoding="utf-8"))
    for finding in report["findings"]:
        assert finding["source"] == "deterministic"


def test_malformed_llm_output_fails_closed_and_does_not_mutate_report(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "w2.xlsx"
    _create(p)
    monkeypatch.chdir(tmp_path)

    audit_result = runner.invoke(app, ["audit", str(p), "--deterministic"])
    assert audit_result.exit_code == 0

    report_path = Path(".sheetproof/sheetproof-report.json")
    before = report_path.read_text(encoding="utf-8")

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {
                "enabled": True,
                "provider": "local",
                "model": "qwen",
                "base_url": "http://localhost:11434",
            }
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.explain_with_ollama",
        lambda prompt, model, base_url="http://localhost:11434": "",
    )

    result = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert result.exit_code != 0

    after = report_path.read_text(encoding="utf-8")
    assert before == after


def test_audit_does_not_call_network(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "w3.xlsx"
    _create(p)
    monkeypatch.chdir(tmp_path)

    called = {"value": False}

    def _fake_urlopen(*args, **kwargs):
        called["value"] = True
        raise AssertionError("Network call should not happen in deterministic audit")

    monkeypatch.setattr("sheetproof.llm.local_explainer.urlopen", _fake_urlopen)

    result = runner.invoke(app, ["audit", str(p), "--deterministic"])
    assert result.exit_code == 0
    assert called["value"] is False
