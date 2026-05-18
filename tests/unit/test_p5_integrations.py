import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=NOW()"
    wb.save(path)


def test_export_integrations_outputs_stable_profiles(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "m.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    res = runner.invoke(app, ["export-integrations"])
    assert res.exit_code == 0
    ci = json.loads(Path(".sheetproof/integrations/ci-annotations.json").read_text(encoding="utf-8"))
    tickets = json.loads(Path(".sheetproof/integrations/ticket-export.json").read_text(encoding="utf-8"))
    assert ci["profile"] == "ci_annotations_v1"
    assert tickets["profile"] == "ticket_export_v1"

    # stability check across repeated exports
    first_ci = Path(".sheetproof/integrations/ci-annotations.json").read_text(encoding="utf-8")
    first_tickets = Path(".sheetproof/integrations/ticket-export.json").read_text(encoding="utf-8")
    assert runner.invoke(app, ["export-integrations"]).exit_code == 0
    second_ci = Path(".sheetproof/integrations/ci-annotations.json").read_text(encoding="utf-8")
    second_tickets = Path(".sheetproof/integrations/ticket-export.json").read_text(encoding="utf-8")
    assert first_ci == second_ci
    assert first_tickets == second_tickets
