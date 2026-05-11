import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    wb.save(path)


def test_structured_explanation_written(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "m.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)

    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {"enabled": True, "provider": "local", "model": "qwen", "base_url": "http://localhost:11434"}
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.explain_with_ollama",
        lambda prompt, model, base_url="http://localhost:11434": json.dumps(
            {
                "summary": "Cell C1 doubles growth input.",
                "risks": ["Input volatility"],
                "reviewer_actions": ["Verify B1 ownership"],
                "citations": [{"cell": "Summary!B1", "reason": "Direct dependency"}],
            }
        ),
    )

    r = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert r.exit_code == 0
    payload = json.loads(Path(".sheetproof/explanations.json").read_text(encoding="utf-8"))
    exp = payload["explanations"][-1]
    assert exp["source"] == "llm_explanation"
    assert "summary" in exp["explanation"]
