from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def test_help_renders() -> None:
    result = runner.invoke(app, ["--help"])
    assert result.exit_code == 0
    assert "spreadsheet audit" in result.stdout.lower()


def test_audit_writes_workbook_index(tmp_path: Path) -> None:
    workbook_path = tmp_path / "audit.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B1"] = "=1+1"
    wb.save(workbook_path)

    result = runner.invoke(app, ["audit", str(workbook_path)])
    assert result.exit_code == 0
    assert "Workbook index written" in result.stdout
    assert Path(".sheetproof/workbook-index.json").exists()
    assert Path(".sheetproof/formula-map.json").exists()
    assert Path(".sheetproof/dependency-graph.json").exists()
    assert Path(".sheetproof/sheetproof-report.md").exists()
    assert Path(".sheetproof/sheetproof-report.json").exists()
    assert Path(".sheetproof/risk-cells.csv").exists()
    assert Path(".sheetproof/assumption-register.csv").exists()
