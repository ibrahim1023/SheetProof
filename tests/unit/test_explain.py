from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Revenue Growth"
    ws["B1"] = 0.12
    ws["C1"] = "=B1*2"
    wb.save(path)


def test_explain_requires_deterministic_artifacts(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "model.xlsx"
    _create_workbook(workbook_path)

    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["explain", str(workbook_path), "--cell", "Summary!C1"])
    assert result.exit_code != 0
    assert not Path(".sheetproof/sheetproof-report.json").exists()


def test_explain_uses_ollama_and_does_not_mutate_findings(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "model.xlsx"
    _create_workbook(workbook_path)

    monkeypatch.chdir(tmp_path)
    audit_result = runner.invoke(app, ["audit", str(workbook_path)])
    assert audit_result.exit_code == 0

    report_path = Path(".sheetproof/sheetproof-report.json")
    original_report = report_path.read_text(encoding="utf-8")

    # Enable local provider in test without modifying real project config.
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "llm": {
                "enabled": True,
                "provider": "local",
                "model": "qwen2.5",
                "base_url": "http://localhost:11434",
            }
        },
    )
    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.explain_with_ollama",
        lambda prompt, model, base_url="http://localhost:11434": "Cell C1 doubles the growth value.",
    )

    explain_result = runner.invoke(app, ["explain", str(workbook_path), "--cell", "Summary!C1"])
    assert explain_result.exit_code == 0
    assert "doubles the growth value" in explain_result.stdout

    after_report = report_path.read_text(encoding="utf-8")
    assert after_report == original_report
