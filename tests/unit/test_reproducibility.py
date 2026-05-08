import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.12
    ws["C1"] = "=B1*2"
    ws["D1"] = "=C1+1"
    wb.save(path)


def test_audit_writes_repro_manifest(tmp_path: Path, monkeypatch) -> None:
    workbook_path = tmp_path / "model.xlsx"
    _create_workbook(workbook_path)

    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["audit", str(workbook_path), "--deterministic"])
    assert result.exit_code == 0

    manifest_path = Path(".sheetproof/reproducibility-manifest.json")
    assert manifest_path.exists()
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert payload["artifacts"]
    assert any(item["file"] == "sheetproof-report.json" for item in payload["artifacts"])


def test_deterministic_mode_repeated_runs_have_same_canonical_hashes(
    tmp_path: Path, monkeypatch
) -> None:
    workbook_path = tmp_path / "model.xlsx"
    _create_workbook(workbook_path)

    monkeypatch.chdir(tmp_path)

    first = runner.invoke(app, ["audit", str(workbook_path), "--deterministic"])
    assert first.exit_code == 0
    first_manifest = json.loads(
        Path(".sheetproof/reproducibility-manifest.json").read_text(encoding="utf-8")
    )

    second = runner.invoke(app, ["audit", str(workbook_path), "--deterministic"])
    assert second.exit_code == 0
    second_manifest = json.loads(
        Path(".sheetproof/reproducibility-manifest.json").read_text(encoding="utf-8")
    )

    first_map = {x["file"]: x["canonical_sha256"] for x in first_manifest["artifacts"]}
    second_map = {x["file"]: x["canonical_sha256"] for x in second_manifest["artifacts"]}
    assert first_map == second_map
