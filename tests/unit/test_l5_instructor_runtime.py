import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    wb.save(path)


def test_explain_with_instructor_engine(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "m.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "local_only": False,
            "observability": {"trace_backend": "local"},
            "llm": {
                "enabled": True,
                "provider": "openai",
                "model": "gpt-5.5",
                "constrained_output_engine": "instructor",
                "max_retries": 1,
                "max_steps": 20,
            },
        },
    )

    class _FakeClient:
        def create(self, response_model, messages, max_retries):  # noqa: ANN001
            return response_model.model_validate(
                {
                    "summary": "Instructor structured output for this cell.",
                    "risks": [],
                    "reviewer_actions": ["Check source assumption."],
                    "citations": [{"cell": "Summary!B1", "reason": "Direct dependency"}],
                }
            )

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.instructor.from_provider",
        lambda *args, **kwargs: _FakeClient(),
    )

    r = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert r.exit_code == 0
    assert "Instructor structured output" in r.stdout
    payload = json.loads(Path(".sheetproof/explanations.json").read_text(encoding="utf-8"))
    assert payload["explanations"][-1]["source"] == "llm_explanation"
