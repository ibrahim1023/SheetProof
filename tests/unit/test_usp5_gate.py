import json
from pathlib import Path
import zipfile

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _make_audit_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=NOW()"
    ws["D1"] = "=[ext.xlsx]Sheet1!A1"
    wb.save(path)


def _make_diff_pair(old_path: Path, new_path: Path) -> None:
    old = Workbook()
    ws = old.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.1
    old.save(old_path)

    new = Workbook()
    ws2 = new.active
    ws2.title = "Inputs"
    ws2["A1"] = "Growth Rate"
    ws2["B1"] = 0.2
    hid = new.create_sheet("HiddenNew")
    hid.sheet_state = "hidden"
    hid["A1"] = 1
    ws2["C1"] = "=[ext.xlsx]Sheet1!A1"
    new.save(new_path)


def test_gate_audit_fails_with_threshold_and_writes_result(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "audit.xlsx"
    _make_audit_workbook(p)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(
        app,
        [
            "gate",
            "--workbook",
            str(p),
            "--max-external-references",
            "0",
        ],
    )
    assert result.exit_code != 0

    gate_path = Path(".sheetproof/gate-result.json")
    assert gate_path.exists()
    payload = json.loads(gate_path.read_text(encoding="utf-8"))
    assert payload["passed"] is False
    assert payload["exit_code"] == 12


def test_gate_diff_fails_hidden_threshold(tmp_path: Path, monkeypatch) -> None:
    old_p = tmp_path / "old.xlsx"
    new_p = tmp_path / "new.xlsx"
    _make_diff_pair(old_p, new_p)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(
        app,
        [
            "gate",
            "--old-workbook",
            str(old_p),
            "--new-workbook",
            str(new_p),
            "--max-new-hidden-sheets",
            "0",
        ],
    )
    assert result.exit_code != 0
    payload = json.loads(Path(".sheetproof/gate-result.json").read_text(encoding="utf-8"))
    assert payload["mode"] == "diff"
    assert any(f["rule"] == "max_new_hidden_sheets" for f in payload["failures"])


def test_gate_passes_when_thresholds_allow(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "audit-ok.xlsx"
    _make_audit_workbook(p)
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(
        app,
        [
            "gate",
            "--workbook",
            str(p),
            "--max-external-references",
            "10",
            "--max-high-risk-findings",
            "10",
        ],
    )
    assert result.exit_code == 0
    payload = json.loads(Path(".sheetproof/gate-result.json").read_text(encoding="utf-8"))
    assert payload["passed"] is True
    assert payload["exit_code"] == 0

    traces = Path(".sheetproof/traces.jsonl").read_text(encoding="utf-8").splitlines()
    assert any('"event": "gate_start"' in line for line in traces)
    assert any('"event": "gate_complete"' in line for line in traces)
    assert any('"request_id":' in line for line in traces)
    assert any('"latency_ms":' in line for line in traces)


def test_gate_fails_on_unattested_feature_threshold(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "audit-macro.xlsx"
    _make_audit_workbook(p)
    with zipfile.ZipFile(p, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"dummy")
    monkeypatch.chdir(tmp_path)

    result = runner.invoke(
        app,
        [
            "gate",
            "--workbook",
            str(p),
            "--max-unattested-features",
            "0",
        ],
    )
    assert result.exit_code != 0
    payload = json.loads(Path(".sheetproof/gate-result.json").read_text(encoding="utf-8"))
    assert any(f["rule"] == "max_unattested_features" for f in payload["failures"])
