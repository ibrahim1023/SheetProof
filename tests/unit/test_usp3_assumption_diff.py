import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _write_old_new(old_path: Path, new_path: Path) -> None:
    old = Workbook()
    ws = old.active
    ws.title = "Inputs"
    ws["A1"] = "Discount Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*100"
    ws["D1"] = "=C1+1"
    old.save(old_path)

    new = Workbook()
    ws2 = new.active
    ws2.title = "Inputs"
    ws2["A1"] = "Discount Rate"
    ws2["B1"] = 0.12
    ws2["C1"] = "=B1*100"
    ws2["D1"] = "=C1+1"
    new.save(new_path)


def test_assumption_register_has_confidence_and_category(tmp_path: Path, monkeypatch) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.2
    ws["C1"] = "=B1*2"
    p = tmp_path / "a.xlsx"
    wb.save(p)

    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["audit", str(p), "--deterministic"])
    assert result.exit_code == 0

    payload = json.loads(Path(".sheetproof/sheetproof-report.json").read_text(encoding="utf-8"))
    assumptions = payload["assumptions"]
    assert assumptions
    assert assumptions[0]["confidence"] in {"high", "medium", "low"}
    assert assumptions[0]["category"]


def test_diff_writes_assumption_diff_with_deltas(tmp_path: Path, monkeypatch) -> None:
    old_path = tmp_path / "old.xlsx"
    new_path = tmp_path / "new.xlsx"
    _write_old_new(old_path, new_path)

    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["diff", str(old_path), str(new_path)])
    assert result.exit_code == 0

    p = Path(".sheetproof/assumption-diff.json")
    assert p.exists()
    payload = json.loads(p.read_text(encoding="utf-8"))
    deltas = payload["assumption_deltas"]
    assert deltas
    assert deltas[0]["old_value"] != deltas[0]["new_value"]
    assert deltas[0]["absolute_change"] is not None
