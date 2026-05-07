from pathlib import Path

from openpyxl import Workbook

from sheetproof.formulas.consistency import detect_formula_consistency_issues
from sheetproof.formulas.extractor import extract_formula_inventory
from sheetproof.graph.builder import build_dependency_graph
from sheetproof.graph.impact import compute_downstream_impact
from sheetproof.workbook.parser import parse_workbook


def test_formula_inventory_and_graph(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    ws["D1"] = "=C1+1"

    p = tmp_path / "f.xlsx"
    wb.save(p)

    idx = parse_workbook(p)
    inv = extract_formula_inventory(idx)
    assert len(inv) == 2
    assert inv[0].references

    graph = build_dependency_graph(inv)
    impact = compute_downstream_impact(graph)
    assert "Inputs!B1" in graph.nodes
    assert impact["Inputs!B1"] >= 1


def test_consistency_issue_detection(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws["B2"] = "=A2+1"
    ws["C2"] = "=A2+1"
    ws["D2"] = "=A2+2"

    p = tmp_path / "consistency.xlsx"
    wb.save(p)

    idx = parse_workbook(p)
    inv = extract_formula_inventory(idx)
    issues = detect_formula_consistency_issues(inv)
    assert any(i.cell == "D2" for i in issues)
