from pathlib import Path
import zipfile

from openpyxl import Workbook
from openpyxl.comments import Comment

from sheetproof.workbook.parser import parse_workbook, write_workbook_index


def test_parse_workbook_extracts_core_fields(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Revenue Growth"
    ws["B1"] = 0.12
    ws["C1"] = "=B1*2"
    ws["A1"].comment = Comment("Assumption cell", "qa")

    hidden = wb.create_sheet("HiddenCalc")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "=Inputs!C1"

    very_hidden = wb.create_sheet("VeryHiddenSheet")
    very_hidden.sheet_state = "veryHidden"

    workbook_path = tmp_path / "model.xlsx"
    wb.save(workbook_path)

    index = parse_workbook(workbook_path)

    assert index.workbook == "model.xlsx"
    assert index.sheet_count == 3
    assert "HiddenCalc" in index.hidden_sheets
    assert "VeryHiddenSheet" in index.very_hidden_sheets

    inputs_sheet = next(s for s in index.sheets if s.name == "Inputs")
    assert inputs_sheet.formula_cells == 1
    assert inputs_sheet.comment_cells == 1
    assert any(c.cell == "C1" and c.formula == "=B1*2" for c in inputs_sheet.cells)


def test_write_workbook_index_creates_json(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1

    workbook_path = tmp_path / "simple.xlsx"
    wb.save(workbook_path)

    index = parse_workbook(workbook_path)
    out_path = write_workbook_index(index, tmp_path / ".sheetproof")

    assert out_path.exists()
    text = out_path.read_text(encoding="utf-8")
    assert "simple.xlsx" in text


def test_parse_workbook_sets_cannot_attest_for_unsupported_features(tmp_path: Path) -> None:
    wb = Workbook()
    workbook_path = tmp_path / "macro_like.xlsx"
    wb.save(workbook_path)

    with zipfile.ZipFile(workbook_path, "a") as zf:
        zf.writestr("xl/vbaProject.bin", b"dummy")

    index = parse_workbook(workbook_path)
    assert index.attestation_status == "cannot_attest"
    assert "UNSUPPORTED:VBA_MACROS" in index.warning_codes
