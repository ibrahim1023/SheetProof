import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app
from sheetproof.config.loader import load_config


runner = CliRunner()


def _wb(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    wb.save(path)


def test_local_only_blocks_hosted_provider(tmp_path: Path, monkeypatch) -> None:
    p = tmp_path / "l.xlsx"
    _wb(p)
    monkeypatch.chdir(tmp_path)
    assert runner.invoke(app, ["audit", str(p), "--deterministic"]).exit_code == 0

    monkeypatch.setattr(
        "sheetproof.llm.local_explainer.load_config",
        lambda config_path=None: {
            "local_only": True,
            "llm": {"enabled": True, "provider": "openai", "model": "gpt-5.5"},
        },
    )
    r = runner.invoke(app, ["explain", str(p), "--cell", "Summary!C1"])
    assert r.exit_code != 0


def test_config_validates_llm_timeout_and_retries(tmp_path: Path) -> None:
    bad = tmp_path / "bad.yml"
    bad.write_text("schema_version: 1\nllm:\n  timeout_seconds: 0\n", encoding="utf-8")
    try:
        load_config(config_path=bad)
        assert False, "Expected ValueError"
    except ValueError as exc:
        assert "timeout_seconds" in str(exc)


def test_eval_min_pass_rate_threshold(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.chdir(tmp_path)
    ds = Path("d.json")
    ds.write_text(
        json.dumps({"cases": [{"id": "bad", "output_json": "{}"}]}),
        encoding="utf-8",
    )
    out = Path("out.json")
    r = runner.invoke(
        app,
        ["eval-explain", "--dataset", str(ds), "--output", str(out), "--min-pass-rate", "0.5"],
    )
    assert r.exit_code == 21
