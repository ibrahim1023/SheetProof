import json
from pathlib import Path

from openpyxl import Workbook
from typer.testing import CliRunner

from sheetproof.cli import app


runner = CliRunner()


def _mk(old_p: Path, new_p: Path) -> None:
    old = Workbook()
    ws = old.active
    ws.title = "Inputs"
    ws["A1"] = "Rate"
    ws["B1"] = 0.1
    ws["C1"] = "=B1*2"
    old.save(old_p)

    new = Workbook()
    ws2 = new.active
    ws2.title = "Inputs"
    ws2["A1"] = "Rate"
    ws2["B1"] = 0.2
    ws2["C1"] = "=B1*3"
    ws2["D1"] = 9
    new.save(new_p)


def test_diff_artifact_order_is_deterministic(tmp_path: Path, monkeypatch) -> None:
    old_p = tmp_path / "old.xlsx"
    new_p = tmp_path / "new.xlsx"
    _mk(old_p, new_p)
    monkeypatch.chdir(tmp_path)

    r1 = runner.invoke(app, ["diff", str(old_p), str(new_p)])
    assert r1.exit_code == 0
    d1 = json.loads(Path(".sheetproof/workbook-diff.json").read_text(encoding="utf-8"))

    r2 = runner.invoke(app, ["diff", str(old_p), str(new_p)])
    assert r2.exit_code == 0
    d2 = json.loads(Path(".sheetproof/workbook-diff.json").read_text(encoding="utf-8"))

    assert d1 == d2
