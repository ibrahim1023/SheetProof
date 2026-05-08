from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from sheetproof.reproducibility import write_stable_json
from sheetproof.workbook.metadata import extract_workbook_metadata
from sheetproof.workbook.models import CellRecord, SheetIndex, WorkbookIndex, utc_now_iso


def _extract_external_links(wb) -> list[str]:
    links: list[str] = []
    rels = getattr(wb, "_external_links", [])
    for link in rels:
        target = getattr(getattr(link, "file_link", None), "Target", None)
        if target:
            links.append(str(target))
    return sorted(set(links))


def parse_workbook(path: Path, deterministic: bool = False) -> WorkbookIndex:
    wb = load_workbook(path, data_only=False)

    sheet_indexes: list[SheetIndex] = []
    hidden_sheets: list[str] = []
    very_hidden_sheets: list[str] = []

    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            hidden_sheets.append(ws.title)
        elif ws.sheet_state == "veryHidden":
            very_hidden_sheets.append(ws.title)

        cells: list[CellRecord] = []
        formula_cells = 0
        comment_cells = 0

        for row in ws.iter_rows():
            for c in row:
                if c.value is None and c.comment is None:
                    continue

                formula: str | None = None
                if isinstance(c.value, str) and c.value.startswith("="):
                    formula = c.value
                    formula_cells += 1

                comment_text = c.comment.text if c.comment else None
                if comment_text:
                    comment_cells += 1

                cells.append(
                    CellRecord(
                        sheet=ws.title,
                        cell=c.coordinate,
                        value=c.value,
                        formula=formula,
                        comment=comment_text,
                    )
                )

        sheet_indexes.append(
            SheetIndex(
                name=ws.title,
                visibility=ws.sheet_state,
                max_row=ws.max_row,
                max_column=ws.max_column,
                populated_cells=len(cells),
                formula_cells=formula_cells,
                comment_cells=comment_cells,
                merged_ranges=[str(rng) for rng in ws.merged_cells.ranges],
                cells=cells,
            )
        )

    defined_names = sorted(list(wb.defined_names.keys()))
    external_links = _extract_external_links(wb)

    return WorkbookIndex(
        workbook=path.name,
        workbook_path=str(path.resolve()),
        generated_at_utc=None if deterministic else utc_now_iso(),
        sheet_count=len(wb.sheetnames),
        sheet_names=wb.sheetnames,
        hidden_sheets=hidden_sheets,
        very_hidden_sheets=very_hidden_sheets,
        defined_names=defined_names,
        external_links=external_links,
        metadata=extract_workbook_metadata(wb),
        sheets=sheet_indexes,
        warnings=[],
    )


def write_workbook_index(index: WorkbookIndex, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "workbook-index.json"
    write_stable_json(out_file, index.to_dict())
    return out_file
